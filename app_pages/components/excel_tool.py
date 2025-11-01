"""
Excel Tool Component - 将截图文件批量插入到Excel文件中
"""
import streamlit as st
import os
import time
import logging
import zipfile
import shutil
from pathlib import Path
from typing import List, Tuple, Optional
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as ExcelImage

# 尝试导入pandas和xlrd用于支持.xls格式
try:
    import pandas as pd
    XLS_SUPPORT = True
except ImportError:
    XLS_SUPPORT = False
    pd = None

# ==================== Excel Tool Helper Classes ====================
# 支持的图片格式
SUPPORTED_IMAGE_FORMATS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.webp'}

# 默认配置
DEFAULT_CONFIG = {
    'header_rows': 3,
    'spacing_rows': 2,
    'image_max_width': 700,
    'image_max_height': 500,
    'output_filename': 'screenshots.xlsx',
    'sheet_name_max_length': 31,
}

def setup_logging(level: str = 'INFO') -> logging.Logger:
    """设置日志记录"""
    logger = logging.getLogger('exceltool')
    logger.setLevel(getattr(logging, level.upper()))
    if not logger.handlers:
        handler = logging.StreamHandler()
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        handler.setFormatter(formatter)
        logger.addHandler(handler)
    return logger

logger = setup_logging()

def cleanup_old_temp_files(temp_base_dir: Path, max_age_hours: int = 24) -> int:
    """
    清理过期的临时文件
    
    Args:
        temp_base_dir: 临时文件基础目录
        max_age_hours: 文件最大保留时间（小时），默认24小时
    
    Returns:
        清理的文件/目录数量
    """
    if not temp_base_dir.exists():
        return 0
    
    import time
    current_time = time.time()
    max_age_seconds = max_age_hours * 3600
    cleaned_count = 0
    
    try:
        for item in temp_base_dir.iterdir():
            try:
                # 跳过清理记录文件
                if item.name == ".last_cleanup_date":
                    continue
                
                # 获取文件/目录的修改时间
                mtime = item.stat().st_mtime
                age = current_time - mtime
                
                # 如果文件超过最大保留时间，则删除
                if age > max_age_seconds:
                    if item.is_dir():
                        shutil.rmtree(item)
                        logger.info(f"清理过期目录: {item}")
                    else:
                        item.unlink()
                        logger.info(f"清理过期文件: {item}")
                    cleaned_count += 1
            except Exception as e:
                logger.warning(f"清理文件失败 {item}: {e}")
                continue
    
    except Exception as e:
        logger.error(f"清理临时文件目录失败: {e}")
    
    return cleaned_count

def check_and_cleanup(temp_base_dir: Path, cleanup_hour: int = 23, cleanup_minute: int = 59, max_age_hours: int = 24) -> bool:
    """
    检查时间并执行清理
    
    Args:
        temp_base_dir: 临时文件基础目录
        cleanup_hour: 清理时间（小时），默认23
        cleanup_minute: 清理时间（分钟），默认59
        max_age_hours: 文件最大保留时间（小时），默认24小时
    
    Returns:
        是否执行了清理
    """
    cleanup_record_file = temp_base_dir / ".last_cleanup_date"
    
    current_time = time.time()
    current_datetime = time.localtime(current_time)
    current_hour = current_datetime.tm_hour
    current_minute = current_datetime.tm_min
    current_date_str = time.strftime("%Y-%m-%d", current_datetime)
    
    # 检查是否到了清理时间（清理时间前后5分钟窗口，确保能执行到）
    time_window = 5  # 5分钟窗口
    
    # 计算时间窗口的开始和结束分钟
    if cleanup_minute >= time_window:
        minute_start = cleanup_minute - time_window
        minute_end = cleanup_minute
        check_hour = cleanup_hour
    else:
        # 如果分钟数小于窗口，检查上一个小时的最后几分钟
        minute_start = 60 - (time_window - cleanup_minute)
        minute_end = cleanup_minute
        check_hour = cleanup_hour
    
    # 检查是否在清理时间窗口内
    in_time_window = False
    
    # 情况1：清理时间在当前小时（例如23:59，窗口是23:54-23:59）
    if cleanup_minute >= time_window:
        if current_hour == cleanup_hour and minute_start <= current_minute <= minute_end:
            in_time_window = True
    # 情况2：清理时间跨越小时边界（例如00:03，窗口是23:58-00:03）
    else:
        if current_hour == cleanup_hour and current_minute <= minute_end:
            in_time_window = True
        elif current_hour == (cleanup_hour - 1) % 24 and current_minute >= minute_start:
            in_time_window = True
    
    if in_time_window:
        # 读取上次清理日期
        last_cleanup_date = None
        if cleanup_record_file.exists():
            try:
                last_cleanup_date = cleanup_record_file.read_text().strip()
            except:
                pass
        
        # 如果今天还没有清理过，则执行清理
        if last_cleanup_date != current_date_str:
            cleaned_count = cleanup_old_temp_files(temp_base_dir, max_age_hours=max_age_hours)
            if cleaned_count > 0:
                logger.info(f"定时清理（{current_date_str} {cleanup_hour:02d}:{cleanup_minute:02d}）：清理了 {cleaned_count} 个过期临时文件/目录")
            
            # 记录本次清理日期
            try:
                cleanup_record_file.write_text(current_date_str)
            except:
                pass
            
            return True
    
    return False

def extract_zip_to_temp_dir(uploaded_zip, temp_base_dir: Path) -> Path:
    """解压ZIP文件到临时目录，返回解压后的目录路径"""
    # 创建唯一的临时目录
    import uuid
    extract_dir = temp_base_dir / f"extracted_{uuid.uuid4().hex[:8]}"
    extract_dir.mkdir(parents=True, exist_ok=True)
    
    # 保存ZIP文件
    zip_path = extract_dir / uploaded_zip.name
    with open(zip_path, "wb") as f:
        f.write(uploaded_zip.getbuffer())
    
    # 解压ZIP文件
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_dir)
    
    # 删除ZIP文件本身
    zip_path.unlink()
    
    return extract_dir

def save_uploaded_files_to_temp_dir(uploaded_files: List, temp_base_dir: Path) -> Path:
    """保存上传的图片文件到临时目录，返回临时目录路径"""
    # 创建唯一的临时目录
    import uuid
    temp_dir = temp_base_dir / f"images_{uuid.uuid4().hex[:8]}"
    temp_dir.mkdir(parents=True, exist_ok=True)
    
    # 保存所有上传的文件
    for uploaded_file in uploaded_files:
        file_path = temp_dir / uploaded_file.name
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
    
    return temp_dir

def get_image_files(directory: Path) -> List[Path]:
    """获取目录中的所有图片文件"""
    image_files = []
    if not directory.exists():
        return image_files
    for file_path in directory.iterdir():
        if file_path.is_file() and file_path.suffix.lower() in SUPPORTED_IMAGE_FORMATS:
            image_files.append(file_path)
    image_files.sort(key=lambda x: x.name)
    return image_files

def get_subdirectories(directory: Path) -> List[Path]:
    """获取目录中的所有子目录"""
    subdirs = []
    if not directory.exists():
        return subdirs
    for item in directory.iterdir():
        if item.is_dir():
            subdirs.append(item)
    # 按目录名排序
    subdirs.sort(key=lambda x: x.name)
    return subdirs

def find_image_folders(root_dir: Path, ignore_dirs: set = None) -> List[Path]:
    """
    递归查找所有包含图片的文件夹（只返回包含图片的叶子文件夹）
    
    Args:
        root_dir: 根目录路径
        ignore_dirs: 要忽略的文件夹名集合（如 _MACOSX, __MACOSX等）
    
    Returns:
        包含图片的文件夹列表（只返回直接包含图片的文件夹，不返回父文件夹）
    """
    if ignore_dirs is None:
        ignore_dirs = {'_macosx', '__macosx', '.ds_store', '.git', '.svn', 'thumbs.db'}
    
    image_folders = []
    
    def _scan_directory(directory: Path):
        """递归扫描目录，找到所有直接包含图片的文件夹"""
        # 跳过系统文件夹
        if directory.name.lower() in ignore_dirs:
            return
        
        # 检查当前目录是否包含图片
        image_files = get_image_files(directory)
        
        # 检查子目录
        subdirs = get_subdirectories(directory)
        has_image_subdirs = False
        
        # 先递归检查所有子目录
        for subdir in subdirs:
            if subdir.name.lower() not in ignore_dirs:
                _scan_directory(subdir)
                # 检查子目录是否包含图片
                if get_image_files(subdir):
                    has_image_subdirs = True
        
        # 如果当前目录包含图片，且没有包含图片的子目录，则添加当前目录
        # 这样可以避免重复：如果父文件夹和子文件夹都有图片，只添加子文件夹
        if image_files and not has_image_subdirs:
            image_folders.append(directory)
    
    _scan_directory(root_dir)
    
    # 按路径排序
    image_folders.sort(key=lambda x: str(x))
    return image_folders

def sanitize_sheet_name(name: str) -> str:
    """清理sheet名称，确保符合Excel要求"""
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '_')
    max_length = DEFAULT_CONFIG['sheet_name_max_length']
    if len(name) > max_length:
        name = name[:max_length]
    if not name.strip():
        name = 'Sheet'
    return name.strip()

class ImageProcessor:
    """图片处理器"""
    def __init__(self, max_width: int = None, max_height: int = None, 
                 fixed_width: int = None, fixed_height: int = None):
        self.max_width = max_width or DEFAULT_CONFIG['image_max_width']
        self.max_height = max_height or DEFAULT_CONFIG['image_max_height']
        self.fixed_width = fixed_width
        self.fixed_height = fixed_height
    
    def resize_image(self, image_path: Path) -> Tuple[ExcelImage, Tuple[int, int]]:
        """调整图片大小并转换为Excel图片对象"""
        import io
        try:
            with Image.open(image_path) as img:
                if img.mode in ('RGBA', 'LA', 'P'):
                    background = Image.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'P':
                        img = img.convert('RGBA')
                    background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                    img = background
                elif img.mode != 'RGB':
                    img = img.convert('RGB')
                
                width, height = img.size
                
                if self.fixed_width and self.fixed_height:
                    new_width = self.fixed_width
                    new_height = self.fixed_height
                    img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                elif self.fixed_width:
                    scale = self.fixed_width / width
                    new_width = self.fixed_width
                    new_height = int(height * scale)
                    img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                elif self.fixed_height:
                    scale = self.fixed_height / height
                    new_width = int(width * scale)
                    new_height = self.fixed_height
                    img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                else:
                    scale_w = self.max_width / width if width > self.max_width else 1
                    scale_h = self.max_height / height if height > self.max_height else 1
                    scale = min(scale_w, scale_h, 1)
                    
                    if scale < 1:
                        new_width = int(width * scale)
                        new_height = int(height * scale)
                        img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                    else:
                        new_width, new_height = width, height
                
                img_buffer = io.BytesIO()
                img.save(img_buffer, format='PNG')
                img_buffer.seek(0)
                
                excel_img = ExcelImage(img_buffer)
                excel_img.width = new_width
                excel_img.height = new_height
                
                return excel_img, (new_width, new_height)
        except Exception as e:
            logger.error(f"处理图片 {image_path} 时出错: {e}")
            raise

class ExcelProcessor:
    """Excel处理器"""
    def __init__(self, header_rows: int = None, spacing_rows: int = None, show_titles: bool = True, respect_header_rows: bool = True):
        self.header_rows = header_rows or DEFAULT_CONFIG['header_rows']
        self.spacing_rows = spacing_rows or DEFAULT_CONFIG['spacing_rows']
        self.show_titles = show_titles
        self.respect_header_rows = respect_header_rows
        self.image_processor = ImageProcessor()
        self.workbook = None
        self.current_row = {}
    
    def create_workbook(self) -> Workbook:
        """创建新的工作簿"""
        self.workbook = Workbook()
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        return self.workbook
    
    def load_workbook(self, file_path: Path) -> Workbook:
        """加载现有的工作簿（支持.xlsx和.xls格式）"""
        try:
            file_ext = file_path.suffix.lower()
            
            # 如果是.xls格式，需要先转换为.xlsx
            if file_ext == '.xls':
                if not XLS_SUPPORT:
                    raise ImportError("需要安装pandas和xlrd库以支持.xls格式文件")
                
                # 使用pandas读取.xls文件
                xls_file = pd.ExcelFile(file_path)
                
                # 创建临时.xlsx文件
                temp_xlsx = file_path.parent / f"{file_path.stem}_temp.xlsx"
                with pd.ExcelWriter(temp_xlsx, engine='openpyxl') as writer:
                    for sheet_name in xls_file.sheet_names:
                        df = pd.read_excel(xls_file, sheet_name=sheet_name)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 使用openpyxl加载转换后的文件
                self.workbook = load_workbook(temp_xlsx)
                
                # 删除临时文件
                try:
                    temp_xlsx.unlink()
                except:
                    pass
            else:
                # 直接使用openpyxl加载.xlsx文件
                self.workbook = load_workbook(file_path)
            
            # 初始化current_row
            for sheet_name in self.workbook.sheetnames:
                if sheet_name not in self.current_row:
                    sheet = self.workbook[sheet_name]
                    max_row = sheet.max_row
                    if self.respect_header_rows:
                        self.current_row[sheet_name] = self.header_rows + 1
                    else:
                        if max_row > 0:
                            for row in range(max_row, max(1, max_row - 5), -1):
                                if any(cell.value for cell in sheet[row]):
                                    self.current_row[sheet_name] = row + 1
                                    break
                            else:
                                self.current_row[sheet_name] = self.header_rows + 1
                        else:
                            self.current_row[sheet_name] = self.header_rows + 1
            return self.workbook
        except Exception as e:
            logger.error(f"加载工作簿失败 {file_path}: {e}")
            raise
    
    def get_or_create_sheet(self, sheet_name: str) -> str:
        """获取或创建工作表"""
        if not self.workbook:
            self.create_workbook()
        
        clean_name = sanitize_sheet_name(sheet_name)
        existing_sheet = self.find_sheet_by_name(clean_name)
        
        if existing_sheet:
            return existing_sheet
        else:
            sheet = self.workbook.create_sheet(title=clean_name)
            self.current_row[clean_name] = self.header_rows + 1
            # 设置列宽：A列用于标题，需要足够宽以显示文件名（60字符宽度）
            sheet.column_dimensions['A'].width = 60
            sheet.column_dimensions['B'].width = 100
            return clean_name
    
    def find_sheet_by_name(self, target_name: str) -> Optional[str]:
        """根据名称查找工作表（忽略大小写）"""
        if not self.workbook:
            return None
        target_lower = target_name.lower()
        for sheet_name in self.workbook.sheetnames:
            if sheet_name.lower() == target_lower:
                return sheet_name
        return None
    
    def add_image_to_sheet(self, sheet_name: str, image_path: Path, image_title: str = None) -> None:
        """向指定工作表添加图片"""
        if not self.workbook:
            logger.error("工作簿不存在")
            return
        
        actual_sheet_name = self.find_sheet_by_name(sheet_name)
        if not actual_sheet_name:
            logger.error(f"工作表 {sheet_name} 不存在")
            return
        
        sheet = self.workbook[actual_sheet_name]
        
        if actual_sheet_name not in self.current_row:
            max_row = sheet.max_row
            if max_row > 0:
                for row in range(max_row, max(1, max_row - 5), -1):
                    if any(cell.value for cell in sheet[row]):
                        self.current_row[actual_sheet_name] = row + 1
                        break
                else:
                    self.current_row[actual_sheet_name] = self.header_rows + 1
            else:
                self.current_row[actual_sheet_name] = self.header_rows + 1
        
        current_row = self.current_row.get(actual_sheet_name, self.header_rows + 1)
        
        try:
            excel_img, (width, height) = self.image_processor.resize_image(image_path)
            
            if self.show_titles and image_title:
                title_cell = sheet.cell(row=current_row, column=1)
                title_cell.value = image_title
                title_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                # 设置标题行高度，确保文件名不被遮挡
                sheet.row_dimensions[current_row].height = 40
                # 确保列宽足够显示文件名
                if sheet.column_dimensions['A'].width < 60:
                    sheet.column_dimensions['A'].width = 60
                current_row += 1
            
            # 图片插入到标题行下方
            excel_img.anchor = f'A{current_row}'
            sheet.add_image(excel_img)
            
            # 设置图片行的行高，确保图片完全显示且不会遮挡下一行的内容
            row_height_points = height * 0.75  # 像素转点
            
            # 根据图片高度动态调整额外空间
            if height < 300:
                extra_space = 30
            elif height < 600:
                extra_space = 40
            else:
                extra_space = 50
            
            row_height = row_height_points + extra_space
            
            # 确保最小行高，避免行高太小
            if row_height < 120:
                row_height = 120
            
            sheet.row_dimensions[current_row].height = row_height
            
            # 更新到下一行时，确保跳过足够的行数
            current_row += 1 + self.spacing_rows
            self.current_row[sheet_name] = current_row
            
        except Exception as e:
            logger.error(f"添加图片失败 {image_path} 到 {sheet_name}: {e}")
    
    def process_directory(self, directory: Path, sheet_name: str = None) -> None:
        """处理目录中的所有图片"""
        if not sheet_name:
            sheet_name = directory.name
        
        actual_sheet_name = self.get_or_create_sheet(sheet_name)
        image_files = get_image_files(directory)
        
        if not image_files:
            logger.warning(f"目录 {directory} 中没有找到图片文件")
            return
        
        for i, image_path in enumerate(image_files, 1):
            image_title = f"{i}. {image_path.stem}"
            self.add_image_to_sheet(actual_sheet_name, image_path, image_title)
    
    def save_workbook(self, output_path: Path) -> None:
        """保存工作簿"""
        if not self.workbook:
            logger.error("没有工作簿可保存")
            return
        try:
            self.workbook.save(output_path)
            logger.info(f"工作簿已保存到: {output_path}")
        except Exception as e:
            logger.error(f"保存工作簿失败 {output_path}: {e}")
            raise


def render_excel_tool():
    """渲染Excel工具界面"""
    st.markdown("## 📊 Excel Tool - Screenshot to Excel")
    
    st.info("将目录中的截图文件批量插入到Excel文件中，支持新建文件或追加到现有文件")
    
    # 使用方法介绍
    with st.expander("📖 详细使用说明（点击展开）", expanded=False):
        st.markdown("""
        ### 📁 文件上传设置
        
        #### 上传方式
        - **类型**: 单选按钮
        - **选项**: 
          - **ZIP文件（推荐 - 保持文件夹结构）**：上传ZIP压缩包，保持文件夹结构，每个文件夹对应一个Sheet（推荐方式）
          - **多个图片文件**：上传多个图片文件，所有图片放在同一个Sheet中
        - **说明**: 
          - ⚠️ **浏览器限制**：浏览器不支持直接选择文件夹，只能逐个选择文件
          - 💡 **推荐使用ZIP文件方式**：可以一次性上传整个文件夹，更加方便快捷
          - ZIP文件方式支持文件夹结构，每个文件夹会创建一个对应的Sheet
        - **ZIP文件说明**: 
          - 上传ZIP文件后会自动解压
          - 如果ZIP中包含多个子文件夹，每个文件夹会创建一个对应的Sheet
          - 文件夹名就是Sheet名
          - 如果ZIP中没有子文件夹，所有图片会放到一个名为"Screenshots"的Sheet中
        - **多个图片文件说明**:
          - ⚠️ 浏览器限制，无法直接选择文件夹，需要逐个选择图片文件
          - 可以同时选择多个图片文件上传（使用Ctrl/Cmd+点击多选）
          - 所有图片会放在同一个名为"Screenshots"的Sheet中
          - 图片会按文件名排序
          - 💡 如果文件较多（>10个），强烈建议使用ZIP文件方式
        
        #### 本地路径（可选）
        - **类型**: 复选框 + 文本输入框
        - **可见性**: 默认隐藏，需要勾选复选框才显示
        - **说明**: 仅用于本地开发测试，在线环境中请使用文件上传功能
        - **注意事项**: ⚠️ 在线环境中无法使用本地路径
        
        #### File Mode（文件模式）
        - **类型**: 复选框
        - **选项**: "Use existing Excel file"（使用已存在的Excel文件）
        - **说明**: 切换文件模式
          - **未勾选**（默认）：创建新Excel文件，处理完成后可直接下载
          - **勾选**：追加到已存在的Excel文件，处理完成后可直接下载
        - **效果**: 切换时会显示/隐藏相应的文件选择选项
        
        #### Existing Excel File（已存在的Excel文件）
        - **类型**: 文件上传器
        - **可见性**: 仅在"使用已存在Excel文件"模式下显示
        - **说明**: 上传要追加截图的Excel文件
        - **操作**: 使用文件上传器上传Excel文件（支持.xlsx和.xls格式）
        - **验证**: 文件必须有效且为.xlsx或.xls格式
        - **注意**: 如果上传.xls格式文件，需要安装pandas和xlrd库（会自动安装）
        
        #### 文件下载
        - **说明**: 处理完成后会自动生成Excel文件并提供下载按钮
        - **新建模式**: 文件名为 `screenshots_{目录名}_{时间戳}.xlsx`
        - **追加模式**: 文件名为 `updated_{原文件名}_{时间戳}.xlsx`
        
        ### 🖼️ 图片设置（Image Settings）
        
        #### Max Width（最大宽度）
        - **类型**: 数字输入框
        - **单位**: 像素
        - **默认值**: `800`
        - **说明**: 图片的最大宽度限制
        - **作用**: 与 Max Height 配合使用，保持图片比例的同时限制最大尺寸
        - **注意事项**: 仅在未设置固定尺寸时生效
        
        #### Max Height（最大高度）
        - **类型**: 数字输入框
        - **单位**: 像素
        - **默认值**: `600`
        - **说明**: 图片的最大高度限制
        - **作用**: 与 Max Width 配合使用，保持图片比例的同时限制最大尺寸
        - **注意事项**: 仅在未设置固定尺寸时生效
        
        #### Fixed Width（固定宽度）
        - **类型**: 复选框 + 数字输入框（可选）
        - **单位**: 像素
        - **默认值**: 未使用（需要勾选复选框才会显示输入框）
        - **说明**: 强制设置所有图片的固定宽度
        - **使用方式**:
          - 只设置 Fixed Width：高度按比例缩放
          - 同时设置 Fixed Width 和 Fixed Height：强制调整为指定尺寸
        - **注意事项**: 与 Max Width 互斥，设置固定宽度时忽略最大宽度
        
        #### Fixed Height（固定高度）
        - **类型**: 复选框 + 数字输入框（可选）
        - **单位**: 像素
        - **默认值**: 未使用（需要勾选复选框才会显示输入框）
        - **说明**: 强制设置所有图片的固定高度
        - **使用方式**:
          - 只设置 Fixed Height：宽度按比例缩放
          - 同时设置 Fixed Width 和 Fixed Height：强制调整为指定尺寸
        - **注意事项**: 与 Max Height 互斥，设置固定高度时忽略最大高度
        
        #### Hide image titles（隐藏图片标题）
        - **类型**: 复选框
        - **默认值**: 未勾选（显示标题）
        - **说明**: 控制是否在图片前显示文件名标题
        - **效果**:
          - **未勾选**：每张图片前显示 "1. 图片名称" 格式的标题
          - **勾选**：只插入图片，不显示标题
        
        ### 📐 布局设置（Layout Settings）
        
        #### Header Rows（Header行数）
        - **类型**: 数字输入框
        - **默认值**: `1`
        - **说明**: Excel工作表顶部的预留行数
        - **作用**: 图片会从第 `header_rows + 1` 行开始插入
        - **使用场景**: 需要在Excel顶部添加表头、说明等信息时
        
        #### Image Spacing（图片间距）
        - **类型**: 数字输入框
        - **单位**: 行数
        - **默认值**: `2`
        - **说明**: 图片之间的空行数
        - **作用**: 控制图片之间的间距，增加可读性
        - **取值范围**: 建议 0-10
        
        ### 🔘 操作按钮
        
        #### Start Processing（开始处理）
        - **功能**: 开始处理截图并生成Excel文件
        - **验证**: 点击前会验证所有输入
        - **状态**: 处理过程中会显示进度条
        - **进度**: 底部状态栏显示处理进度
        - **完成**: 处理完成后会显示成功信息并提供下载按钮
        
        #### Clear（清空）
        - **功能**: 清空所有输入，恢复默认值
        - **操作**: 清空所有字段和选择
        
        #### Exit（退出）
        - **功能**: 关闭程序
        
        ### ⚠️ 注意事项
        
        - 需要安装 `openpyxl` 和 `Pillow` 库：`pip install openpyxl Pillow`
        - 如果要支持.xls格式，还需要安装：`pip install pandas xlrd`
        - 图片文件会按文件名排序后插入
        - 处理大文件时可能需要一些时间，请耐心等待
        - 如果输出目录不存在，会自动创建
        - 支持多文件夹处理：如果输入路径下有多个子文件夹，每个文件夹会创建一个对应的Sheet
        """)
    
    st.markdown("---")
    
    # 初始化session state
    if 'last_input_path' not in st.session_state:
        st.session_state.last_input_path = ""
    if 'last_excel_path' not in st.session_state:
        st.session_state.last_excel_path = ""
    if 'uploaded_files_dir' not in st.session_state:
        st.session_state.uploaded_files_dir = None
    
    # 文件上传设置
    st.markdown("### 📁 文件上传设置")
    
    # 上传方式选择
    upload_mode = st.radio(
        "上传方式",
        ["📦 ZIP文件（推荐 - 保持文件夹结构）", "🖼️ 多个图片文件"],
        help="💡 推荐使用ZIP文件方式：可以一次性上传整个文件夹，保持文件夹结构，每个文件夹对应一个Sheet"
    )
    
    # 创建临时目录用于存储上传的文件
    temp_base_dir = Path(os.path.join(os.path.expanduser("~"), ".streamlit_temp"))
    temp_base_dir.mkdir(parents=True, exist_ok=True)
    
    # 定期清理过期临时文件（默认每天23:59清理，保留24小时内的文件）
    check_and_cleanup(temp_base_dir, cleanup_hour=23, cleanup_minute=59, max_age_hours=24)
    
    uploaded_files_path = None
    
    if upload_mode == "📦 ZIP文件（推荐 - 保持文件夹结构）":
        uploaded_zip = st.file_uploader(
            "上传ZIP文件",
            type=["zip"],
            help="上传包含图片文件的ZIP压缩包，文件夹结构会被保留（每个文件夹对应一个Sheet）",
            key="upload_zip"
        )
        
        if uploaded_zip:
            try:
                # 解压ZIP文件
                with st.spinner("正在解压ZIP文件..."):
                    extracted_dir = extract_zip_to_temp_dir(uploaded_zip, temp_base_dir)
                    st.session_state.uploaded_files_dir = str(extracted_dir)
                    uploaded_files_path = extracted_dir
                    st.success(f"✅ ZIP文件已解压: {uploaded_zip.name}")
                    
                    # 显示文件夹结构预览
                    # 使用智能查找函数，找到所有包含图片的文件夹
                    image_folders = find_image_folders(extracted_dir)
                    
                    if image_folders:
                        st.info(f"📁 检测到 {len(image_folders)} 个包含图片的文件夹，将创建对应的Sheet")
                        with st.expander("查看文件夹结构"):
                            for folder in image_folders:
                                image_count = len(get_image_files(folder))
                                # 显示相对路径，更清晰
                                relative_path = folder.relative_to(extracted_dir)
                                st.write(f"- {relative_path}: {image_count} 张图片")
                    else:
                        # 如果没有找到包含图片的文件夹，检查根目录是否有图片
                        image_count = len(get_image_files(extracted_dir))
                        if image_count > 0:
                            st.info(f"📁 检测到 {image_count} 张图片，将放入 'Screenshots' Sheet")
                        else:
                            st.warning("⚠️ 未找到图片文件，请检查ZIP文件内容")
            except Exception as e:
                st.error(f"❌ 解压ZIP文件失败: {str(e)}")
                uploaded_files_path = None
    
    else:  # 多个图片文件
        st.info("💡 **提示**: 浏览器不支持直接选择文件夹。如果需要上传多个文件，建议使用ZIP文件方式（上方选项），可以一次性上传整个文件夹。")
        
        uploaded_images = st.file_uploader(
            "上传图片文件（可多选）",
            type=["png", "jpg", "jpeg", "gif", "bmp", "webp", "tiff"],
            accept_multiple_files=True,
            help="⚠️ 注意：浏览器限制，无法直接选择文件夹。如果需要上传多个文件，建议使用ZIP文件方式。当前方式需要逐个选择图片文件。",
            key="upload_images"
        )
        
        if uploaded_images:
            if len(uploaded_images) > 10:
                st.warning(f"⚠️ 您选择了 {len(uploaded_images)} 个文件。如果文件较多，建议使用ZIP文件方式上传，更加方便快捷。")
            
            try:
                # 保存上传的图片文件
                logger.info(f"开始保存 {len(uploaded_images)} 张图片文件")
                with st.spinner("正在保存图片文件..."):
                    images_dir = save_uploaded_files_to_temp_dir(uploaded_images, temp_base_dir)
                    st.session_state.uploaded_files_dir = str(images_dir)
                    uploaded_files_path = images_dir
                    st.success(f"✅ 已上传 {len(uploaded_images)} 张图片")
                    logger.info(f"图片文件保存成功: {len(uploaded_images)} 张图片 -> {images_dir}")
            except Exception as e:
                st.error(f"❌ 保存图片文件失败: {str(e)}")
                logger.error(f"保存图片文件失败: {e}", exc_info=True)
                uploaded_files_path = None
    
    # 文件路径设置（保留原有功能，但标记为可选）
    st.markdown("---")
    st.markdown("### 📁 文件路径设置（可选 - 仅用于本地测试）")
    
    use_local_path = st.checkbox(
        "使用本地路径（仅用于本地开发测试）",
        value=False,
        help="⚠️ 在线环境中请使用文件上传功能，此选项仅用于本地开发测试"
    )
    
    input_path = None
    if use_local_path:
        input_path = st.text_input(
            "Input Path（输入路径）",
            value=st.session_state.last_input_path,
            help="选择包含截图文件的目录，可以直接输入路径（仅用于本地测试）",
            key="excel_input_path"
        )
        
        if input_path and os.path.isdir(input_path):
            st.session_state.last_input_path = input_path
            uploaded_files_path = None  # 使用本地路径时，不使用上传的文件
    else:
        # 如果使用上传的文件，使用上传的文件路径
        if st.session_state.uploaded_files_dir:
            input_path = st.session_state.uploaded_files_dir
            uploaded_files_path = Path(input_path)
    
    # 文件模式
    use_existing_file = st.checkbox(
        "File Mode: Use existing Excel file（文件模式：使用已存在的Excel文件）",
        value=False,
        help="未勾选：创建新Excel文件（处理完成后可下载）；勾选：追加到已存在的Excel文件"
    )
    
    # 已存在文件设置（仅在追加模式下显示）
    if use_existing_file:
        uploaded_file = st.file_uploader(
            "Existing Excel File（上传已存在的Excel文件）",
            type=["xlsx", "xls"],
            help="上传要追加截图的Excel文件（支持.xlsx和.xls格式）",
            key="upload_excel"
        )
        
        if uploaded_file:
            temp_dir = os.path.join(os.path.expanduser("~"), ".streamlit_temp")
            os.makedirs(temp_dir, exist_ok=True)
            temp_path = os.path.join(temp_dir, uploaded_file.name)
            with open(temp_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            st.session_state.excel_existing_file = temp_path
            st.info(f"📁 已上传文件: {uploaded_file.name}")
            
            # 如果是.xls格式，提示需要安装依赖
            if uploaded_file.name.lower().endswith('.xls') and not XLS_SUPPORT:
                st.warning("⚠️ 检测到.xls格式文件，需要安装pandas和xlrd库。请运行: pip install pandas xlrd")
    else:
        st.info("ℹ️ 将创建新的Excel文件，处理完成后可直接下载")
    
    st.markdown("---")
    
    # 图片设置
    st.markdown("### 🖼️ 图片设置")
    
    col1, col2 = st.columns(2)
    
    with col1:
        max_width = st.number_input(
            "Max Width（最大宽度，像素）",
            min_value=1,
            value=700,
            help="图片的最大宽度限制，与Max Height配合使用，保持图片比例的同时限制最大尺寸"
        )
        
        use_fixed_width = st.checkbox("使用固定宽度", value=False, key="use_fixed_width")
        fixed_width = None
        if use_fixed_width:
            fixed_width = st.number_input(
                "Fixed Width（固定宽度，像素）",
                min_value=1,
                value=800,
                help="强制设置所有图片的固定宽度。只设置Fixed Width：高度按比例缩放；同时设置Fixed Width和Fixed Height：强制调整为指定尺寸。",
                key="fixed_width_input"
            )
    
    with col2:
        max_height = st.number_input(
            "Max Height（最大高度，像素）",
            min_value=1,
            value=500,
            help="图片的最大高度限制，与Max Width配合使用，保持图片比例的同时限制最大尺寸"
        )
        
        use_fixed_height = st.checkbox("使用固定高度", value=False, key="use_fixed_height")
        fixed_height = None
        if use_fixed_height:
            fixed_height = st.number_input(
                "Fixed Height（固定高度，像素）",
                min_value=1,
                value=600,
                help="强制设置所有图片的固定高度。只设置Fixed Height：宽度按比例缩放；同时设置Fixed Width和Fixed Height：强制调整为指定尺寸。",
                key="fixed_height_input"
            )
    
    hide_image_titles = st.checkbox(
        "Hide image titles（隐藏图片标题）",
        value=False,
        help="勾选后只插入图片，不显示文件名标题"
    )
    
    st.markdown("---")
    
    # 布局设置
    st.markdown("### 📐 布局设置")
    
    col1, col2 = st.columns(2)
    
    with col1:
        header_rows = st.number_input(
            "Header Rows（Header行数）",
            min_value=0,
            value=3,
            help="Excel工作表顶部的预留行数，图片会从第header_rows+1行开始插入"
        )
    
    with col2:
        image_spacing = st.number_input(
            "Image Spacing（图片间距，行数）",
            min_value=0,
            value=2,
            help="图片之间的空行数，控制图片之间的间距"
        )
    
    st.markdown("---")
    
    # 操作按钮
    col1, col2, col3 = st.columns([2, 1, 1])
    
    with col1:
        process_button = st.button("🚀 Start Processing（开始处理）", type="primary", use_container_width=True)
    
    with col2:
        clear_button = st.button("🗑️ Clear（清空）", use_container_width=True)
    
    with col3:
        exit_button = st.button("🚪 Exit（退出）", use_container_width=True)
    
    # 清空按钮逻辑
    if clear_button:
        logger.info("用户点击清空按钮，开始清理临时文件和状态")
        
        # 清理上传的文件目录
        if st.session_state.get('uploaded_files_dir'):
            try:
                temp_path = Path(st.session_state.uploaded_files_dir)
                if temp_path.exists():
                    shutil.rmtree(temp_path)
                    logger.info(f"已删除上传的文件目录: {temp_path}")
            except Exception as e:
                logger.warning(f"删除上传文件目录失败: {e}")
        
        # 清理上传的Excel文件
        if st.session_state.get('excel_existing_file'):
            try:
                excel_file = Path(st.session_state.excel_existing_file)
                if excel_file.exists():
                    excel_file.unlink()
                    logger.info(f"已删除上传的Excel文件: {excel_file}")
            except Exception as e:
                logger.warning(f"删除Excel文件失败: {e}")
        
        # 清理生成的输出文件
        if st.session_state.get('last_output_file'):
            try:
                output_file = Path(st.session_state.last_output_file)
                if output_file.exists():
                    output_file.unlink()
                    logger.info(f"已删除生成的输出文件: {output_file}")
            except Exception as e:
                logger.warning(f"删除输出文件失败: {e}")
        
        # 清空session state
        st.session_state.excel_input_path = ""
        st.session_state.excel_existing_file = ""
        st.session_state.uploaded_files_dir = None
        st.session_state.last_output_file = None
        
        logger.info("清空操作完成")
        st.rerun()
    
    # 退出按钮逻辑
    if exit_button:
        logger.info("用户点击退出按钮，开始清理临时文件和状态")
        
        # 清理上传的文件目录
        if st.session_state.get('uploaded_files_dir'):
            try:
                temp_path = Path(st.session_state.uploaded_files_dir)
                if temp_path.exists():
                    shutil.rmtree(temp_path)
                    logger.info(f"已删除上传的文件目录: {temp_path}")
            except Exception as e:
                logger.warning(f"删除上传文件目录失败: {e}")
        
        # 清理上传的Excel文件
        if st.session_state.get('excel_existing_file'):
            try:
                excel_file = Path(st.session_state.excel_existing_file)
                if excel_file.exists():
                    excel_file.unlink()
                    logger.info(f"已删除上传的Excel文件: {excel_file}")
            except Exception as e:
                logger.warning(f"删除Excel文件失败: {e}")
        
        # 清理生成的输出文件
        if st.session_state.get('last_output_file'):
            try:
                output_file = Path(st.session_state.last_output_file)
                if output_file.exists():
                    output_file.unlink()
                    logger.info(f"已删除生成的输出文件: {output_file}")
            except Exception as e:
                logger.warning(f"删除输出文件失败: {e}")
        
        logger.info("退出操作完成")
        st.stop()
    
    # 处理按钮逻辑
    if process_button:
        errors = []
        
        # 验证输入路径
        if not use_local_path:
            # 使用上传的文件
            if not uploaded_files_path or not uploaded_files_path.exists():
                errors.append("❌ 请先上传文件（ZIP文件或图片文件）")
        else:
            # 使用本地路径
            if not input_path or not os.path.isdir(input_path):
                errors.append("❌ 请输入有效的输入路径（目录）")
            else:
                uploaded_files_path = Path(input_path)
        
        if use_existing_file:
            existing_file = st.session_state.get('excel_existing_file', "")
            if not existing_file or not os.path.isfile(existing_file):
                errors.append("❌ 请上传已存在的Excel文件")
            elif not existing_file.lower().endswith(('.xlsx', '.xls')):
                errors.append("❌ Excel文件必须是.xlsx或.xls格式")
        
        if errors:
            for error in errors:
                st.error(error)
        else:
            try:
                # 确定要处理的路径
                if uploaded_files_path:
                    input_path_obj = Path(uploaded_files_path)
                else:
                    input_path_obj = Path(input_path)
                
                # 创建Excel处理器
                processor = ExcelProcessor(
                    header_rows=header_rows,
                    spacing_rows=image_spacing,
                    show_titles=not hide_image_titles,
                    respect_header_rows=True
                )
                
                # 设置图片处理器参数
                processor.image_processor.max_width = max_width
                processor.image_processor.max_height = max_height
                processor.image_processor.fixed_width = fixed_width
                processor.image_processor.fixed_height = fixed_height
                
                # 处理Excel文件
                if use_existing_file:
                    existing_file = st.session_state.get('excel_existing_file', "")
                    processor.load_workbook(Path(existing_file))
                else:
                    processor.create_workbook()
                
                # 检查是否有包含图片的文件夹（智能查找）
                image_folders = find_image_folders(input_path_obj)
                
                if image_folders:
                    # 有包含图片的文件夹：为每个文件夹创建一个sheet
                    total_images = 0
                    processed_images = 0
                    
                    # 先统计总图片数
                    for folder in image_folders:
                        image_files = get_image_files(folder)
                        total_images += len(image_files)
                    
                    if total_images == 0:
                        st.warning("⚠️ 在文件夹中未找到图片文件（支持：png, jpg, jpeg, gif, bmp, webp）")
                    else:
                        # 进度条
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        # 处理每个包含图片的文件夹
                        for folder in image_folders:
                            # 使用文件夹的相对路径作为sheet名，如果有嵌套则使用最后的文件夹名
                            sheet_name = folder.name
                            image_files = get_image_files(folder)
                            
                            if image_files:
                                status_text.text(f"处理文件夹: {sheet_name} ({len(image_files)} 张图片)")
                                logger.info(f"开始处理文件夹: {sheet_name}, 包含 {len(image_files)} 张图片")
                                
                                # 获取或创建工作表（使用文件夹名作为sheet名）
                                actual_sheet_name = processor.get_or_create_sheet(sheet_name)
                                logger.info(f"处理Sheet: {actual_sheet_name}, 包含 {len(image_files)} 张图片")
                                
                                # 逐个处理图片
                                for idx, image_path in enumerate(image_files, 1):
                                    status_text.text(f"处理中: {sheet_name}/{image_path.name} ({processed_images + idx}/{total_images})")
                                    image_title = f"{idx}. {image_path.stem}"
                                    logger.debug(f"添加图片到Sheet: {actual_sheet_name}, 图片: {image_path.name}")
                                    processor.add_image_to_sheet(actual_sheet_name, image_path, image_title)
                                    progress_bar.progress((processed_images + idx) / total_images)
                                
                                logger.info(f"Sheet {actual_sheet_name} 处理完成，共处理 {len(image_files)} 张图片")
                                processed_images += len(image_files)
                        
                        # 保存文件到临时目录
                        temp_dir = os.path.join(os.path.expanduser("~"), ".streamlit_temp")
                        os.makedirs(temp_dir, exist_ok=True)
                        
                        if use_existing_file:
                            # 追加模式：保存到新的临时文件（不覆盖原文件）
                            output_filename = f"updated_{Path(existing_file).stem}_{int(time.time())}.xlsx"
                        else:
                            # 新建模式：使用输入目录名作为文件名
                            if use_local_path:
                                output_filename = f"screenshots_{input_path_obj.name}_{int(time.time())}.xlsx"
                            else:
                                output_filename = f"screenshots_{int(time.time())}.xlsx"
                        
                        output_path = Path(temp_dir) / output_filename
                        
                        status_text.text("正在保存Excel文件...")
                        logger.info(f"保存Excel文件到: {output_path}")
                        processor.save_workbook(output_path)
                        logger.info(f"Excel文件保存成功: {output_path}")
                        
                        # 保存到session state
                        if use_local_path:
                            st.session_state.last_input_path = str(input_path_obj)
                        # 保存输出文件路径供下载使用
                        st.session_state.last_output_file = str(output_path)
                        
                        # 清理临时文件（如果不是本地路径）
                        if not use_local_path and uploaded_files_path and uploaded_files_path.exists():
                            try:
                                shutil.rmtree(uploaded_files_path)
                                logger.info(f"处理完成，已删除上传的文件目录: {uploaded_files_path}")
                            except Exception as e:
                                logger.warning(f"删除上传文件目录失败: {e}")
                        
                        # 完成提示
                        progress_bar.progress(1.0)
                        status_text.empty()
                        st.success(f"✅ 处理完成！共处理 {len(image_folders)} 个Sheet（对应 {len(image_folders)} 个文件夹），{total_images} 张图片。")
                        
                        # 提供下载链接
                        logger.info(f"准备下载Excel文件: {output_filename} (路径: {output_path})")
                        with open(output_path, 'rb') as f:
                            file_data = f.read()
                            download_clicked = st.download_button(
                                label="📥 下载Excel文件",
                                data=file_data,
                                file_name=output_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_{int(time.time())}"
                            )
                            
                            # 如果下载按钮被点击，延迟删除文件（给用户时间下载）
                            if download_clicked:
                                logger.info(f"用户点击下载按钮，文件: {output_filename}")
                                # 延迟删除文件（60秒后）
                                import threading
                                def delayed_delete(file_path, delay=60):
                                    """延迟删除文件（60秒后）"""
                                    time.sleep(delay)
                                    try:
                                        if Path(file_path).exists():
                                            Path(file_path).unlink()
                                            logger.info(f"延迟删除输出文件: {file_path}")
                                    except Exception as e:
                                        logger.warning(f"延迟删除文件失败 {file_path}: {e}")
                                
                                thread = threading.Thread(target=delayed_delete, args=(output_path,))
                                thread.daemon = True
                                thread.start()
                else:
                    # 没有子文件夹：处理当前目录的图片
                    image_files = get_image_files(input_path_obj)
                    
                    if not image_files:
                        st.warning("⚠️ 在指定目录中未找到图片文件（支持：png, jpg, jpeg, gif, bmp, webp）")
                    else:
                        # 进度条
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        # 获取或创建工作表
                        actual_sheet_name = processor.get_or_create_sheet("Screenshots")
                        logger.info(f"处理Sheet: {actual_sheet_name}, 包含 {len(image_files)} 张图片")
                        
                        # 逐个处理图片并显示进度
                        for idx, image_path in enumerate(image_files, 1):
                            status_text.text(f"处理中: {image_path.name} ({idx}/{len(image_files)})")
                            image_title = f"{idx}. {image_path.stem}"
                            logger.debug(f"添加图片到Sheet: {actual_sheet_name}, 图片: {image_path.name}")
                            processor.add_image_to_sheet(actual_sheet_name, image_path, image_title)
                            progress_bar.progress(idx / len(image_files))
                        
                        logger.info(f"Sheet {actual_sheet_name} 处理完成，共处理 {len(image_files)} 张图片")
                        
                        # 保存文件到临时目录
                        temp_dir = os.path.join(os.path.expanduser("~"), ".streamlit_temp")
                        os.makedirs(temp_dir, exist_ok=True)
                        
                        if use_existing_file:
                            # 追加模式：保存到新的临时文件（不覆盖原文件）
                            output_filename = f"updated_{Path(existing_file).stem}_{int(time.time())}.xlsx"
                        else:
                            # 新建模式：使用输入目录名作为文件名
                            if use_local_path:
                                output_filename = f"screenshots_{input_path_obj.name}_{int(time.time())}.xlsx"
                            else:
                                output_filename = f"screenshots_{int(time.time())}.xlsx"
                        
                        output_path = Path(temp_dir) / output_filename
                        
                        status_text.text("正在保存Excel文件...")
                        logger.info(f"保存Excel文件到: {output_path}")
                        processor.save_workbook(output_path)
                        logger.info(f"Excel文件保存成功: {output_path}")
                        
                        # 保存到session state
                        if use_local_path:
                            st.session_state.last_input_path = str(input_path_obj)
                        # 保存输出文件路径供下载使用
                        st.session_state.last_output_file = str(output_path)
                        
                        # 清理临时文件（如果不是本地路径）
                        if not use_local_path and uploaded_files_path and uploaded_files_path.exists():
                            try:
                                shutil.rmtree(uploaded_files_path)
                                logger.info(f"处理完成，已删除上传的文件目录: {uploaded_files_path}")
                            except Exception as e:
                                logger.warning(f"删除上传文件目录失败: {e}")
                        
                        # 完成提示
                        progress_bar.progress(1.0)
                        status_text.empty()
                        st.success(f"✅ 处理完成！共处理 {len(image_files)} 张图片。")
                        
                        # 提供下载链接
                        logger.info(f"准备下载Excel文件: {output_filename} (路径: {output_path})")
                        with open(output_path, 'rb') as f:
                            file_data = f.read()
                            download_clicked = st.download_button(
                                label="📥 下载Excel文件",
                                data=file_data,
                                file_name=output_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_{int(time.time())}"
                            )
                            
                            # 如果下载按钮被点击，标记文件待删除
                            # 注意：由于Streamlit的限制，无法直接检测下载完成，所以会在下次清空或退出时删除
                            if download_clicked:
                                logger.info(f"用户点击下载按钮，文件: {output_filename}")
                                # 延迟删除文件（给用户时间下载）
                                import threading
                                def delayed_delete(file_path, delay=60):
                                    """延迟删除文件（60秒后）"""
                                    time.sleep(delay)
                                    try:
                                        if Path(file_path).exists():
                                            Path(file_path).unlink()
                                            logger.info(f"延迟删除输出文件: {file_path}")
                                    except Exception as e:
                                        logger.warning(f"延迟删除文件失败 {file_path}: {e}")
                                
                                thread = threading.Thread(target=delayed_delete, args=(output_path,))
                                thread.daemon = True
                                thread.start()
                    
            except Exception as e:
                st.error(f"❌ 处理过程中出错: {str(e)}")
                import traceback
                st.code(traceback.format_exc())

